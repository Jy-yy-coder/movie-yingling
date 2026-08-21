# -*- coding: utf-8 -*-
"""按「第一语言」对电影分类导出 Excel。
- 新增「第一语言」列（取 languages 用 | 分隔后的第一个）
- 全部数据放在「全部(按语言分组)」sheet：同一第一语言的排在一起，
  语言按该语言电影数量从多到少排列，组内按评分从高到低
- 每种第一语言再单独生成一个 sheet，便于逐语言查看
"""
import sys
sys.stdout.reconfigure(encoding="utf-8")
import re
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from crawler_movie import MOVIES_CSV, CSV_FIELDS

OUT_XLSX = MOVIES_CSV.parent / "movies_by_language.xlsx"


def clean_sheet_name(name):
    """Excel sheet 名不能含 []:*?/\\ 且 <=31 字符"""
    name = re.sub(r"[\[\]:*?/\\]", "", name)
    return name[:31] if name else "未知"


def main():
    df = pd.read_csv(MOVIES_CSV, dtype=str).fillna("")
    # 第一语言
    df["第一语言"] = df["languages"].str.split("|").str[0].str.strip().replace("", "未知")
    # 评分转数值用于排序
    df["_r"] = pd.to_numeric(df["rating"], errors="coerce").fillna(0)

    # 语言按数量降序排序
    order = df["第一语言"].value_counts()
    df["_lang_cnt"] = df["第一语言"].map(order)
    df_sorted = df.sort_values(
        by=["_lang_cnt", "第一语言", "_r"],
        ascending=[False, True, False],
    ).drop(columns=["_r", "_lang_cnt"])

    # 输出列顺序：第一语言 放到 languages 后面
    cols = CSV_FIELDS.copy()
    idx = cols.index("languages") + 1
    out_cols = cols[:idx] + ["第一语言"] + cols[idx:]

    with pd.ExcelWriter(OUT_XLSX, engine="openpyxl") as writer:
        # 主表
        df_sorted[out_cols].to_excel(writer, sheet_name="全部(按语言分组)", index=False)
        # 各语言分表
        for lang in order.index:
            sub = df_sorted[df_sorted["第一语言"] == lang][out_cols]
            sheet = clean_sheet_name(f"{lang}({len(sub)})")
            sub.to_excel(writer, sheet_name=sheet, index=False)

        # 统一美化
        header_fill = PatternFill("solid", fgColor="4472C4")
        header_font = Font(color="FFFFFF", bold=True)
        for ws in writer.book.worksheets:
            # 表头样式
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(vertical="center")
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions
            # 列宽
            widths = {"movie_id": 10, "poster_url": 16, "title": 34, "year": 6,
                      "director": 18, "writer": 18, "actors": 30, "genres": 16,
                      "countries": 14, "languages": 20, "第一语言": 10,
                      "runtime": 12, "rating": 6, "rating_count": 10, "summary": 60}
            for i, col in enumerate(out_cols, 1):
                ws.column_dimensions[get_column_letter(i)].width = widths.get(col, 14)

    # 控制台汇总
    print(f"已导出: {OUT_XLSX}")
    print(f"总计 {len(df)} 部，{order.size} 种第一语言，各语言 sheet 已生成")
    print("\n语言分布(按数量):")
    for lang, cnt in order.items():
        print(f"  {lang:<8} {cnt} 部")


if __name__ == "__main__":
    main()
